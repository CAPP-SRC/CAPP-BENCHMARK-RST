#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║         CNC OPERATION SHEET — MULTI-GROUP VENDOR RATING             ║
║                                                                      ║
║  Confronta N operation sheet PDF presenti in una cartella            ║
║  e restituisce una classifica complessiva stile Vendor Rating.       ║
║                                                                      ║
║  Uso:  python multi_benchmark_cnc.py  <cartella_pdf>                 ║
║        python multi_benchmark_cnc.py  a.pdf b.pdf c.pdf              ║
║                                                                      ║
║  Opzioni:                                                            ║
║    --xlsx  <file.xlsx>   Esporta risultati in Excel                  ║
║    --tool-life <minuti>  Soglia vita utile utensile (default: 20)    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import argparse
import re
import sys
import os
from collections import defaultdict
from pathlib import Path
from itertools import combinations

try:
    import pdfplumber
except ImportError:
    sys.exit("Errore: installa pdfplumber con  pip install pdfplumber")


# ═══════════════════════════════════════════════════════════════════
# 1. PDF PARSER  (identico a benchmark_cnc.py)
# ═══════════════════════════════════════════════════════════════════

def parse_cycle_time(text: str) -> int:
    text = text.strip().split("(")[0].strip()
    h = m = s = 0
    hm = re.search(r'(\d+)h', text)
    mm = re.search(r'(\d+)m', text)
    sm = re.search(r'(\d+)s', text)
    if hm: h = int(hm.group(1))
    if mm: m = int(mm.group(1))
    if sm: s = int(sm.group(1))
    return h * 3600 + m * 60 + s


def extract_field(text: str, field: str, as_float: bool = False):
    pattern = rf'{field}:\s*([\d.,]+)'
    match = re.search(pattern, text)
    if match:
        val = match.group(1).replace(",", "")
        return float(val) if as_float else val
    return None


def detect_strategy(op_text: str) -> str:
    strat_match = re.search(r'Strategy:\s*([A-Za-z]+(?:\s+[A-Za-z0-9]+)?)', op_text)
    if strat_match:
        raw = strat_match.group(1).strip()
        known = ["Adaptive", "Facing", "Contour 2D", "Contour", "Drilling",
                 "Scallop", "Bore", "Pocket", "Slot", "Trace", "Radial",
                 "Spiral", "Morphed Spiral", "Parallel", "Pencil", "Steep and Shallow"]
        for k in known:
            if raw.startswith(k):
                return k
        return raw.split()[0]
    desc_match = re.search(r'Description:\s*(?:\d+\s+)?(\w+)', op_text)
    if desc_match and desc_match.group(1).lower().startswith("flat"):
        return "Flat"
    return "Unknown"


def extract_product_code(op_text: str) -> str:
    match = re.search(r'Product:\s*(.+?)(?:\n|$)', op_text)
    if match:
        product = match.group(1).strip()
        product = re.split(r'\s{2,}', product)[0].strip()
        product = re.sub(r'^fresa a punta tonda\s*', '', product, flags=re.IGNORECASE)
        product = re.split(r'\s+con\s+inserto', product, flags=re.IGNORECASE)[0].strip()
        return product
    return "N/A"


def extract_short_name(full_name: str, filename: str = "") -> str:
    """Estrae un nome breve dal Document Path o dal nome file."""
    # Cerca pattern NC01, NC02, TP01, TP02, GR01, ecc. nel doc path o nel filename
    for source in [full_name, filename]:
        match = re.search(r'((?:NC|TP|GR)\d+)', source, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    # Fallback: prime parole significative dal doc path
    clean = re.sub(r'[_\-]', ' ', full_name).split()
    return clean[0] if clean else full_name


def parse_pdf(pdf_path: str) -> dict:
    result = {'name': '', 'setups': [], 'path': pdf_path}
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    doc_match = re.search(r'Document Path:\s*(.+)', full_text)
    result['name'] = doc_match.group(1).strip() if doc_match else Path(pdf_path).stem

    setup_blocks = re.split(r'(?=Setup Sheet for Program \d+)', full_text)
    setup_blocks = [b for b in setup_blocks if b.strip() and 'Setup Sheet for Program' in b]

    for block in setup_blocks:
        setup = {'program': '', 'cycle_time_s': 0, 'n_operations': 0, 'n_tools': 0, 'operations': []}
        prog_match = re.search(r'Setup Sheet for Program (\d+)', block)
        if prog_match: setup['program'] = prog_match.group(1)
        nops_match = re.search(r'Number Of Operations:\s*(\d+)', block)
        if nops_match: setup['n_operations'] = int(nops_match.group(1))
        ntools_match = re.search(r'Number Of Tools:\s*(\d+)', block)
        if ntools_match: setup['n_tools'] = int(ntools_match.group(1))
        ct_match = re.search(r'Estimated Cycle Time:\s*([\dhms:]+)', block)
        if ct_match: setup['cycle_time_s'] = parse_cycle_time(ct_match.group(1))

        op_pattern = r'(Operation\s+(\d+)/(\d+)\s+(T\d+)\s+D\d+\s+L\d+.*?)(?=Operation\s+\d+/\d+|$)'
        ops = re.findall(op_pattern, block, re.DOTALL)
        for op_text, op_num, op_total, tool_t in ops:
            cutting = extract_field(op_text, 'Cutting Distance', as_float=True) or 0.0
            rapid = extract_field(op_text, 'Rapid Distance', as_float=True) or 0.0
            feedrate = extract_field(op_text, 'Maximum Feedrate', as_float=True) or 0.0
            op_ct_match = re.search(r'Estimated Cycle Time:\s*([\dhms:]+(?:\s*\([^)]*\))?)', op_text)
            op_ct = parse_cycle_time(op_ct_match.group(1)) if op_ct_match else 0
            desc_match = re.search(r'Description:\s*(.+?)(?:\s{2,}|Maximum|Minimum|$)', op_text)
            description = desc_match.group(1).strip() if desc_match else ""
            strategy = detect_strategy(op_text)
            product = extract_product_code(op_text)
            setup['operations'].append({
                'op_num': int(op_num), 'op_total': int(op_total),
                'description': description, 'strategy': strategy,
                'tool_t': tool_t, 'product': product,
                'cutting_dist': cutting, 'rapid_dist': rapid,
                'max_feedrate': feedrate, 'cycle_time_s': op_ct,
            })
        result['setups'].append(setup)
    return result


# ═══════════════════════════════════════════════════════════════════
# 2. CALCOLO METRICHE
# ═══════════════════════════════════════════════════════════════════

def compute_metrics(parsed: dict, tool_life_s: int = 1200) -> dict:
    all_ops = []
    for setup in parsed['setups']:
        all_ops.extend(setup['operations'])
    if not all_ops:
        print(f"  ⚠ Attenzione: nessuna operazione trovata in '{parsed['name']}', gruppo ignorato.")
        return None

    total_time = sum(s['cycle_time_s'] for s in parsed['setups'])
    if total_time == 0:
        total_time = sum(o['cycle_time_s'] for o in all_ops)
    total_cut = sum(o['cutting_dist'] for o in all_ops)
    total_rapid = sum(o['rapid_dist'] for o in all_ops)
    n_ops = len(all_ops)
    setup_times = [s['cycle_time_s'] for s in parsed['setups']]
    products = set(o['product'] for o in all_ops if o['product'] != 'N/A')

    tool_changes = 0
    for setup in parsed['setups']:
        ops = setup['operations']
        for i in range(1, len(ops)):
            if ops[i]['tool_t'] != ops[i - 1]['tool_t']:
                tool_changes += 1

    strategies = set(o['strategy'] for o in all_ops)
    strat_time = defaultdict(int)
    strat_count = defaultdict(int)
    tool_time = defaultdict(int)
    tool_trefs = defaultdict(set)
    for o in all_ops:
        strat_time[o['strategy']] += o['cycle_time_s']
        strat_count[o['strategy']] += 1
        tool_time[o['product']] += o['cycle_time_s']
        tool_trefs[o['product']].add(o['tool_t'])

    weighted_feed = sum(o['max_feedrate'] * o['cutting_dist'] for o in all_ops) / total_cut if total_cut else 0
    max_tool_time = max(tool_time.values()) if tool_time else 0
    max_tool_prod = max(tool_time, key=tool_time.get) if tool_time else "N/A"
    n_products = len(products)
    limit = tool_life_s

    short_name = extract_short_name(parsed['name'], Path(parsed.get('path', '')).stem)
    return {
        'group': short_name, 'full_name': parsed['name'],
        'total_time': total_time, 'setup_times': setup_times,
        'total_cut': total_cut, 'total_rapid': total_rapid,
        'n_ops': n_ops,
        'n_ops_per_setup': [len(s['operations']) for s in parsed['setups']],
        'n_products': n_products, 'tc_total': tool_changes,
        'n_strategies': len(strategies), 'strategies': strategies,
        'strat_time': dict(strat_time), 'strat_count': dict(strat_count),
        'tool_time': dict(tool_time),
        'tool_trefs': {k: sorted(v) for k, v in tool_trefs.items()},
        'weighted_feed': weighted_feed,
        'max_tool_time': max_tool_time, 'max_tool_prod': max_tool_prod,
        'tools_over_50': sum(1 for t in tool_time.values() if t / limit > 0.5),
        'tools_over_75': sum(1 for t in tool_time.values() if t / limit > 0.75),
        'tools_over_100': sum(1 for t in tool_time.values() if t / limit > 1.0),
        'avg_util': sum(t / limit for t in tool_time.values()) / len(tool_time) if tool_time else 0,
        'cut_ratio': total_cut / (total_cut + total_rapid) if (total_cut + total_rapid) else 0,
        'ops_per_tool': n_ops / n_products if n_products else 0,
        'productivity': total_cut / (total_time / 60) if total_time else 0,
        'max_tool_pct_cycle': max_tool_time / total_time if total_time else 0,
        'tool_life_s': tool_life_s,
    }


# ═══════════════════════════════════════════════════════════════════
# 3. SISTEMA DI SCORING (VENDOR RATING) — MULTI-GROUP
# ═══════════════════════════════════════════════════════════════════

CATEGORY_WEIGHTS = {
    'Efficienza Temporale': 0.30,
    'Utilizzo Utensili': 0.20,
    'Vita Utile': 0.20,
    'Efficienza di Percorso': 0.15,
    'Complessità del Ciclo': 0.10,
    'Aggressività di Taglio': 0.05,
}


def relative_score_multi(values: list, lower_is_better: bool = True) -> list:
    """
    Punteggio relativo per N gruppi: il migliore prende 100, gli altri in proporzione.
    """
    if all(v == 0 for v in values):
        return [100.0] * len(values)
    if lower_is_better:
        best = min(v for v in values if v > 0) if any(v > 0 for v in values) else 1
        return [round(best / v * 100, 1) if v > 0 else 100.0 for v in values]
    else:
        best = max(values)
        if best == 0:
            return [100.0] * len(values)
        return [round(v / best * 100, 1) for v in values]


def tool_life_score(metrics: dict) -> float:
    limit = metrics['tool_life_s']
    scores = []
    for t in metrics['tool_time'].values():
        pct = t / limit
        if pct <= 0.5: s = 100
        elif pct <= 0.75: s = 80
        elif pct <= 1.0: s = 60
        else: s = max(0, 60 - (pct - 1.0) * 200)
        scores.append(s)
    return round(sum(scores) / len(scores), 1) if scores else 100


def compute_all_scores(metrics_list: list):
    """
    Calcola punteggi per N gruppi simultaneamente.

    Returns:
        drivers: lista di (categoria, nome_driver, [raw_values], [scores], [display_values])
        cat_scores: lista di dict {categoria: score} per ogni gruppo
        totals: lista di float punteggi finali
    """
    N = len(metrics_list)
    drivers = []

    def add(cat, name, raws, scores, displays):
        drivers.append((cat, name, raws, scores, displays))

    # 1. EFFICIENZA TEMPORALE
    vals = [m['total_time'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Efficienza Temporale', 'Tempo ciclo complessivo', vals, scores,
        [fmt_time(v) for v in vals])

    vals = [m['total_time'] / m['n_ops'] if m['n_ops'] else 0 for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Efficienza Temporale', 'Tempo medio per operazione', vals, scores,
        [fmt_time(v) for v in vals])

    # 2. UTILIZZO UTENSILI
    vals = [m['n_products'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Utilizzo Utensili', 'N° utensili univoci', vals, scores,
        [str(v) for v in vals])

    vals = [m['tc_total'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Utilizzo Utensili', 'N° cambi utensile', vals, scores,
        [str(v) for v in vals])

    # 3. VITA UTILE
    tls = [tool_life_score(m) for m in metrics_list]
    add('Vita Utile', 'Score vita utile (non lineare)', tls, tls,
        [f"{v:.1f}/100" for v in tls])

    vals = [m['max_tool_pct_cycle'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Vita Utile', 'Concentrazione utensile più impiegato', vals, scores,
        [f"{v * 100:.1f}%" for v in vals])

    penalties = [max(0, 100 - m['tools_over_100'] * 50) for m in metrics_list]
    add('Vita Utile', 'Penalità superamento vita (−50pt/utensile)',
        [m['tools_over_100'] for m in metrics_list], penalties,
        [f"{m['tools_over_100']} utensili" for m in metrics_list])

    # 4. EFFICIENZA DI PERCORSO
    vals = [m['cut_ratio'] for m in metrics_list]
    scores = relative_score_multi(vals, False)
    add('Efficienza di Percorso', 'Rapporto taglio / (taglio + rapido)', vals, scores,
        [f"{v * 100:.1f}%" for v in vals])

    vals = [m['total_cut'] + m['total_rapid'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Efficienza di Percorso', 'Distanza complessiva', vals, scores,
        [f"{v:.0f} mm" for v in vals])

    # 5. COMPLESSITA'
    vals = [m['n_ops'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Complessità del Ciclo', 'N° operazioni totali', vals, scores,
        [str(v) for v in vals])

    vals = [m['ops_per_tool'] for m in metrics_list]
    scores = relative_score_multi(vals, True)
    add('Complessità del Ciclo', 'Rapporto operazioni / utensile', vals, scores,
        [f"{v:.1f}" for v in vals])

    # 6. AGGRESSIVITA'
    vals = [m['weighted_feed'] for m in metrics_list]
    scores = relative_score_multi(vals, False)
    add('Aggressività di Taglio', 'Feedrate medio ponderato', vals, scores,
        [f"{v:.0f} mm/min" for v in vals])

    vals = [m['productivity'] for m in metrics_list]
    scores = relative_score_multi(vals, False)
    add('Aggressività di Taglio', 'Produttività [mm taglio / min ciclo]', vals, scores,
        [f"{v:.0f}" for v in vals])

    # Calcolo punteggi categoria
    cat_scores = [{} for _ in range(N)]
    for cat in CATEGORY_WEIGHTS:
        cd = [d for d in drivers if d[0] == cat]
        if cd:
            for i in range(N):
                cat_scores[i][cat] = round(sum(d[3][i] for d in cd) / len(cd), 1)

    totals = [round(sum(cs[c] * w for c, w in CATEGORY_WEIGHTS.items()), 1) for cs in cat_scores]

    return drivers, cat_scores, totals


# ═══════════════════════════════════════════════════════════════════
# 4. FORMATTAZIONE OUTPUT
# ═══════════════════════════════════════════════════════════════════

def fmt_time(s: float) -> str:
    s = int(s)
    m, sec = divmod(s, 60)
    h, m = divmod(m, 60)
    return f"{h}h {m:02d}m {sec:02d}s" if h > 0 else f"{m}m {sec:02d}s"


def print_multi_report(metrics_list, drivers, cat_scores, totals):
    N = len(metrics_list)
    names = [m['group'] for m in metrics_list]

    # Classifica per score totale
    ranking = sorted(range(N), key=lambda i: totals[i], reverse=True)

    col_w = max(12, max(len(n) for n in names) + 2)
    W = 40 + col_w * N

    print("\n" + "═" * W)
    print(f"{'VENDOR RATING — MULTI-GROUP BENCHMARK CNC':^{W}}")
    print(f"{N} gruppi confrontati".center(W))
    print("═" * W)

    # ── CLASSIFICA ──
    print(f"\n  {'CLASSIFICA FINALE':^30}")
    print(f"  {'─' * 30}")
    for pos, idx in enumerate(ranking, 1):
        medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(pos, "  ")
        print(f"  {medal} {pos}°  {names[idx]:<16}  {totals[idx]:>6.1f} / 100")

    # ── DETTAGLIO CATEGORIE ──
    print(f"\n  {'CATEGORIA':<32} {'Peso':>6}", end="")
    for i in ranking:
        print(f"  {names[i]:>{col_w}}", end="")
    print()
    print(f"  {'─' * 32} {'─' * 6}", end="")
    for _ in range(N):
        print(f"  {'─' * col_w}", end="")
    print()

    for cat, weight in CATEGORY_WEIGHTS.items():
        scores_cat = [cat_scores[i][cat] for i in range(N)]
        best_val = max(scores_cat)
        print(f"  {cat:<32} {weight * 100:>5.0f}%", end="")
        for i in ranking:
            marker = " ◄" if cat_scores[i][cat] == best_val and scores_cat.count(best_val) == 1 else "  "
            print(f"  {cat_scores[i][cat]:>{col_w - 2}.1f}{marker}", end="")
        print()

    # Totale
    print(f"  {'─' * 32} {'─' * 6}", end="")
    for _ in range(N):
        print(f"  {'─' * col_w}", end="")
    print()
    print(f"  {'TOTALE PESATO':<32} {'100%':>6}", end="")
    for i in ranking:
        marker = " ◄" if totals[i] == max(totals) and totals.count(max(totals)) == 1 else "  "
        print(f"  {totals[i]:>{col_w - 2}.1f}{marker}", end="")
    print()

    # ── DETTAGLIO DRIVER ──
    print(f"\n{'─' * W}")
    print(f"  {'DRIVER':<44}", end="")
    for i in ranking:
        print(f"  {names[i]:>{col_w}}", end="")
    print(f"  {'Best':>{col_w}}")
    print(f"  {'─' * 44}", end="")
    for _ in range(N + 1):
        print(f"  {'─' * col_w}", end="")
    print()

    current_cat = ""
    for cat, driver_name, raws, scores, displays in drivers:
        if cat != current_cat:
            print(f"\n  ▸ {cat} ({CATEGORY_WEIGHTS[cat] * 100:.0f}%)")
            current_cat = cat

        best_score = max(scores)
        best_idx = [i for i in range(N) if scores[i] == best_score]

        print(f"    {driver_name:<42}", end="")
        for i in ranking:
            marker = " ◄" if scores[i] == best_score and len(best_idx) == 1 else "  "
            # Show value and score
            val_str = f"{displays[i]} ({scores[i]:.0f})"
            print(f"  {val_str:>{col_w}}", end="")
        # Best column
        if len(best_idx) == 1:
            print(f"  {names[best_idx[0]]:>{col_w}}", end="")
        else:
            print(f"  {'=':>{col_w}}", end="")
        print()

    # ── ALLARMI VITA UTILE ──
    has_alarms = any(m['tools_over_100'] > 0 for m in metrics_list)
    if has_alarms:
        limit = metrics_list[0]['tool_life_s']
        print(f"\n{'─' * W}")
        print(f"  ⚠  ALLARMI VITA UTILE (soglia {limit // 60} min)")
        for m in metrics_list:
            for p, t in m['tool_time'].items():
                if t > limit:
                    trefs = ", ".join(m['tool_trefs'].get(p, []))
                    print(f"    [{m['group']}]  {p} ({trefs}): {fmt_time(t)} = {t / limit * 100:.1f}% vita")

    # ── METODOLOGIA ──
    print(f"\n{'─' * W}")
    print("  METODOLOGIA")
    print("  • Punteggio relativo: il migliore su ciascun driver ottiene 100, gli altri in proporzione")
    print("  • Vita utile non lineare: ≤50%→100 | 50–75%→80 | 75–100%→60 | >100%→penalità rapida")
    print("  • Penalità assoluta: −50pt per ogni utensile oltre il 100% vita utile")
    print("  • Pesi: Tempo 30% | Utensili 20% | Vita 20% | Percorso 15% | Complessità 10% | Taglio 5%")
    print("═" * W + "\n")


# ═══════════════════════════════════════════════════════════════════
# 5. ESPORTAZIONE EXCEL
# ═══════════════════════════════════════════════════════════════════

def export_multi_xlsx(metrics_list, drivers, cat_scores, totals, xlsx_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Attenzione: openpyxl non installato, export Excel saltato.")
        return

    N = len(metrics_list)
    names = [m['group'] for m in metrics_list]
    ranking = sorted(range(N), key=lambda i: totals[i], reverse=True)

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hfill = PatternFill("solid", fgColor="2F5496")
    cat_font = Font(bold=True, name="Arial", size=10, color="2F5496")
    cat_fill = PatternFill("solid", fgColor="D6E4F0")
    df = Font(name="Arial", size=10)
    bf = Font(bold=True, name="Arial", size=10)
    bf12 = Font(bold=True, name="Arial", size=12)
    bf14 = Font(bold=True, name="Arial", size=14)
    green_font = Font(bold=True, name="Arial", size=10, color="217346")
    red_font = Font(bold=True, name="Arial", size=10, color="C00000")
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    better_fill = PatternFill("solid", fgColor="E2EFDA")
    worse_fill = PatternFill("solid", fgColor="FCE4EC")
    gold_fill = PatternFill("solid", fgColor="FFD700")
    silver_fill = PatternFill("solid", fgColor="E0E0E0")
    bronze_fill = PatternFill("solid", fgColor="F4D3A0")
    medal_fills = [gold_fill, silver_fill, bronze_fill]

    # ═══════════════════ FOGLIO 1: CLASSIFICA ═══════════════════
    ws = wb.active
    ws.title = "Classifica"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + N)
    ws.cell(row=1, column=1, value=f"VENDOR RATING — {N} GRUPPI A CONFRONTO").font = Font(bold=True, name="Arial", size=14, color="2F5496")

    # Podio
    row = 3
    headers_r = ["Pos.", "Gruppo", "Score"]
    for c, h in enumerate(headers_r, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = thin
    row += 1
    for pos, idx in enumerate(ranking, 1):
        ws.cell(row=row, column=1, value=f"{pos}°").font = bf; ws.cell(row=row, column=1).alignment = ca; ws.cell(row=row, column=1).border = thin
        ws.cell(row=row, column=2, value=names[idx]).font = bf12; ws.cell(row=row, column=2).alignment = ca; ws.cell(row=row, column=2).border = thin
        score_cell = ws.cell(row=row, column=3, value=f"{totals[idx]:.1f}")
        score_cell.font = bf14; score_cell.alignment = ca; score_cell.border = thin
        if pos <= 3:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = medal_fills[pos - 1]
        row += 1

    # Dettaglio categorie
    row += 1
    ws.cell(row=row, column=1, value="DETTAGLIO PER CATEGORIA").font = Font(bold=True, name="Arial", size=12, color="2F5496")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + N)
    row += 1

    cat_headers = ["Categoria", "Peso"] + [names[i] for i in ranking]
    for c, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = thin
    row += 1

    for cat, weight in CATEGORY_WEIGHTS.items():
        ws.cell(row=row, column=1, value=cat).font = bf; ws.cell(row=row, column=1).alignment = la; ws.cell(row=row, column=1).border = thin
        ws.cell(row=row, column=2, value=f"{weight * 100:.0f}%").font = df; ws.cell(row=row, column=2).alignment = ca; ws.cell(row=row, column=2).border = thin
        scores_cat = [cat_scores[i][cat] for i in ranking]
        best_cat = max(scores_cat)
        for j, idx in enumerate(ranking):
            cell = ws.cell(row=row, column=3 + j, value=f"{cat_scores[idx][cat]:.1f}")
            cell.font = bf; cell.alignment = ca; cell.border = thin
            if cat_scores[idx][cat] == best_cat and scores_cat.count(best_cat) == 1:
                cell.fill = better_fill
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTALE PESATO").font = bf12; ws.cell(row=row, column=1).fill = cat_fill; ws.cell(row=row, column=1).border = thin
    ws.cell(row=row, column=2, value="100%").font = bf; ws.cell(row=row, column=2).alignment = ca; ws.cell(row=row, column=2).fill = cat_fill; ws.cell(row=row, column=2).border = thin
    for j, idx in enumerate(ranking):
        cell = ws.cell(row=row, column=3 + j, value=f"{totals[idx]:.1f}")
        cell.font = bf14; cell.alignment = ca; cell.border = thin; cell.fill = cat_fill
        if idx == ranking[0]:
            cell.fill = gold_fill

    # Methodology note
    row += 2
    notes = [
        "METODOLOGIA",
        "• Punteggio relativo: il migliore su ciascun driver ottiene 100, gli altri in proporzione.",
        "• Vita utile — scoring non lineare: ≤50% → 100pt | 50–75% → 80pt | 75–100% → 60pt | >100% → penalità rapida verso 0.",
        "• Penalità assoluta: −50 punti per ogni utensile che supera il 100% della vita utile.",
        "• Pesi: Efficienza Temporale 30% | Utilizzo Utensili 20% | Vita Utile 20% | Eff. Percorso 15% | Complessità 10% | Aggressività 5%.",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=note).font = Font(name="Arial", size=9, italic=(not note.startswith("MET")),
                                                            bold=note.startswith("MET"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + N)
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    for j in range(N):
        ws.column_dimensions[get_column_letter(3 + j)].width = 16

    # ═══════════════════ FOGLIO 2: SCORECARD COMPLETA ═══════════════════
    ws2 = wb.create_sheet("Scorecard Dettaglio")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + 2 * N)
    ws2.cell(row=1, column=1, value="SCORECARD — DETTAGLIO DRIVER").font = Font(bold=True, name="Arial", size=13, color="2F5496")

    row = 3
    # Headers: Categoria | Driver | [Valore G1 | Score G1] * N | Best
    h2 = ["", "Driver"]
    for i in ranking:
        h2 += [f"Valore {names[i]}", f"Score {names[i]}"]
    h2.append("Migliore")
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = thin
    row += 1

    current_cat = ""
    for cat, driver_name, raws, scores, displays in drivers:
        if cat != current_cat:
            ncols = 3 + 2 * N
            for c in range(1, ncols + 1):
                ws2.cell(row=row, column=c).fill = cat_fill; ws2.cell(row=row, column=c).border = thin
            ws2.cell(row=row, column=1, value=f"{cat} ({CATEGORY_WEIGHTS[cat] * 100:.0f}%)").font = cat_font
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            ws2.cell(row=row, column=1).fill = cat_fill
            current_cat = cat; row += 1

        ws2.cell(row=row, column=2, value=driver_name).font = df; ws2.cell(row=row, column=2).alignment = la
        best_score = max(scores)
        best_indices = [i for i in range(N) if scores[i] == best_score]

        for j, idx in enumerate(ranking):
            col_val = 3 + 2 * j
            col_sc = 4 + 2 * j
            ws2.cell(row=row, column=col_val, value=displays[idx]).font = df; ws2.cell(row=row, column=col_val).alignment = ca
            sc_cell = ws2.cell(row=row, column=col_sc, value=f"{scores[idx]:.1f}")
            sc_cell.font = bf; sc_cell.alignment = ca
            if scores[idx] == best_score and len(best_indices) == 1:
                sc_cell.fill = better_fill

        # Best column
        if len(best_indices) == 1:
            ws2.cell(row=row, column=3 + 2 * N, value=names[best_indices[0]]).font = green_font
        else:
            ws2.cell(row=row, column=3 + 2 * N, value="=").font = df
        ws2.cell(row=row, column=3 + 2 * N).alignment = ca

        ncols = 3 + 2 * N
        for c in range(1, ncols + 1):
            ws2.cell(row=row, column=c).border = thin
        row += 1

    ws2.column_dimensions['A'].width = 4
    ws2.column_dimensions['B'].width = 44
    for j in range(2 * N + 1):
        ws2.column_dimensions[get_column_letter(3 + j)].width = 16

    # ═══════════════════ FOGLIO 3: VITA UTILE ═══════════════════
    ws3 = wb.create_sheet("Vita Utile")
    all_prods = sorted(set(p for m in metrics_list for p in m['tool_time'].keys()))

    h3 = ["#", "Codice PRODUCT"]
    for i in ranking:
        h3 += [f"Tempo {names[i]}", f"% Vita {names[i]}", f"Stato {names[i]}"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = thin
    ws3.row_dimensions[1].height = 32

    limit = metrics_list[0]['tool_life_s']
    for i, prod in enumerate(all_prods, 1):
        r = i + 1
        ws3.cell(row=r, column=1, value=i).font = df; ws3.cell(row=r, column=1).alignment = ca; ws3.cell(row=r, column=1).border = thin
        ws3.cell(row=r, column=2, value=prod).font = bf; ws3.cell(row=r, column=2).border = thin

        for j, idx in enumerate(ranking):
            col_t = 3 + 3 * j
            col_p = 4 + 3 * j
            col_s = 5 + 3 * j
            m = metrics_list[idx]
            if prod in m['tool_time']:
                ts = m['tool_time'][prod]; pct = ts / limit
                ws3.cell(row=r, column=col_t, value=fmt_time(ts)).font = df; ws3.cell(row=r, column=col_t).alignment = ca
                ws3.cell(row=r, column=col_p, value=f"{pct * 100:.1f}%").font = df; ws3.cell(row=r, column=col_p).alignment = ca
                if pct > 1.0:
                    ws3.cell(row=r, column=col_s, value="⚠ SUPERATO").font = red_font
                    ws3.cell(row=r, column=col_p).fill = worse_fill; ws3.cell(row=r, column=col_p).font = red_font
                elif pct > 0.75:
                    ws3.cell(row=r, column=col_s, value="Attenzione").font = Font(name="Arial", size=10, color="FF8C00")
                    ws3.cell(row=r, column=col_p).fill = PatternFill("solid", fgColor="FFF2CC")
                elif pct > 0.5:
                    ws3.cell(row=r, column=col_s, value="Moderato").font = df
                else:
                    ws3.cell(row=r, column=col_s, value="OK").font = green_font
                    ws3.cell(row=r, column=col_p).fill = better_fill
                ws3.cell(row=r, column=col_s).alignment = ca
            else:
                for cc in [col_t, col_p, col_s]:
                    ws3.cell(row=r, column=cc, value="—").font = Font(name="Arial", size=10, color="AAAAAA")
                    ws3.cell(row=r, column=cc).alignment = ca
            for cc in [col_t, col_p, col_s]:
                ws3.cell(row=r, column=cc).border = thin

    ws3.column_dimensions['A'].width = 4
    ws3.column_dimensions['B'].width = 28
    for j in range(3 * N):
        ws3.column_dimensions[get_column_letter(3 + j)].width = 14

    # ═══════════════════ FOGLIO 4: RADAR DATA ═══════════════════
    ws4 = wb.create_sheet("Dati Radar")
    ws4.cell(row=1, column=1, value="Dati per grafico radar — punteggi per categoria").font = Font(bold=True, name="Arial", size=11, color="2F5496")
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + N)

    row = 3
    h4 = ["Categoria"] + [names[i] for i in ranking]
    for c, h in enumerate(h4, 1):
        cell = ws4.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = thin
    row += 1
    for cat in CATEGORY_WEIGHTS:
        ws4.cell(row=row, column=1, value=cat).font = bf; ws4.cell(row=row, column=1).border = thin
        for j, idx in enumerate(ranking):
            cell = ws4.cell(row=row, column=2 + j, value=cat_scores[idx][cat])
            cell.font = df; cell.alignment = ca; cell.border = thin
        row += 1

    ws4.column_dimensions['A'].width = 28
    for j in range(N):
        ws4.column_dimensions[get_column_letter(2 + j)].width = 14

    wb.save(xlsx_path)
    print(f"\n  ✓ Report Excel salvato in: {xlsx_path}")


# ═══════════════════════════════════════════════════════════════════
# 6. MAIN
# ═══════════════════════════════════════════════════════════════════

def collect_pdfs(inputs: list) -> list:
    """Raccoglie tutti i PDF da una lista di file e/o cartelle."""
    pdfs = []
    for inp in inputs:
        p = Path(inp)
        if p.is_dir():
            found = sorted(p.glob("*.pdf")) + sorted(p.glob("*.PDF"))
            pdfs.extend(found)
        elif p.is_file() and p.suffix.lower() == '.pdf':
            pdfs.append(p)
        else:
            print(f"  ⚠ Ignorato: {inp} (non è un file PDF né una cartella)")
    # Rimuovi duplicati mantenendo ordine
    seen = set()
    unique = []
    for p in pdfs:
        rp = p.resolve()
        if rp not in seen:
            seen.add(rp)
            unique.append(p)
    return unique


def main():
    parser = argparse.ArgumentParser(
        description="CNC Operation Sheet — Multi-Group Vendor Rating Benchmark",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python multi_benchmark_cnc.py  ./pdf_folder/
  python multi_benchmark_cnc.py  NC01.pdf NC02.pdf NC03.pdf TP01.pdf TP02.pdf TP03.pdf
  python multi_benchmark_cnc.py  ./pdf_folder/ --xlsx classifica.xlsx
  python multi_benchmark_cnc.py  ./pdf_folder/ --xlsx classifica.xlsx --tool-life 15
        """)
    parser.add_argument('inputs', nargs='+',
                        help='Uno o più file PDF, oppure una cartella contenente i PDF')
    parser.add_argument('--xlsx', help='Esporta risultati in file Excel', default=None)
    parser.add_argument('--tool-life', type=int, default=20,
                        help='Soglia vita utile utensile in minuti (default: 20)')

    args = parser.parse_args()
    tool_life_s = args.tool_life * 60

    # Raccolta PDF
    pdfs = collect_pdfs(args.inputs)
    if len(pdfs) < 2:
        sys.exit(f"Errore: servono almeno 2 file PDF. Trovati: {len(pdfs)}")

    print(f"\n  Trovati {len(pdfs)} file PDF:")
    for p in pdfs:
        print(f"    • {p.name}")

    # Parsing
    print()
    metrics_list = []
    for pdf_path in pdfs:
        print(f"  Parsing {pdf_path.name} ...")
        parsed = parse_pdf(str(pdf_path))
        n_ops = sum(len(s['operations']) for s in parsed['setups'])
        print(f"  → {parsed['name']}: {n_ops} operazioni in {len(parsed['setups'])} setup")
        m = compute_metrics(parsed, tool_life_s)
        if m is not None:
            metrics_list.append(m)

    if len(metrics_list) < 2:
        sys.exit(f"Errore: servono almeno 2 gruppi validi. Parsati con successo: {len(metrics_list)}")

    # Check nomi duplicati
    group_names = [m['group'] for m in metrics_list]
    if len(set(group_names)) != len(group_names):
        # Aggiungi suffisso progressivo ai duplicati
        counts = defaultdict(int)
        for m in metrics_list:
            counts[m['group']] += 1
        if any(c > 1 for c in counts.values()):
            seen = defaultdict(int)
            for m in metrics_list:
                if counts[m['group']] > 1:
                    seen[m['group']] += 1
                    m['group'] = f"{m['group']}_{seen[m['group']]}"

    # Scoring
    drivers, cat_scores, totals = compute_all_scores(metrics_list)

    # Output
    print_multi_report(metrics_list, drivers, cat_scores, totals)

    # Excel
    if args.xlsx:
        export_multi_xlsx(metrics_list, drivers, cat_scores, totals, args.xlsx)


if __name__ == '__main__':
    main()
