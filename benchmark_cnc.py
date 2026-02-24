#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║              CNC OPERATION SHEET — VENDOR RATING BENCHMARK          ║
║                                                                      ║
║  Confronta 2 operation sheet PDF (Fusion 360 / HSMWorks Setup Sheet) ║
║  e restituisce un punteggio 0–100 stile Vendor Rating.               ║
║                                                                      ║
║  Uso:  python benchmark_cnc.py  <pdf_gruppo_A>  <pdf_gruppo_B>       ║
║                                                                      ║
║  Opzioni:                                                            ║
║    --xlsx  <file.xlsx>   Esporta risultati in Excel                  ║
║    --tool-life <minuti>  Soglia vita utile utensile (default: 20)    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    sys.exit("Errore: installa pdfplumber con  pip install pdfplumber")


# ═══════════════════════════════════════════════════════════════════
# 1. PDF PARSER
# ═══════════════════════════════════════════════════════════════════

def parse_cycle_time(text: str) -> int:
    """Converte stringhe come '4m:39s', '26s', '1h:02m:30s' in secondi."""
    text = text.strip().split("(")[0].strip()
    h = m = s = 0
    hm = re.search(r'(\d+)h', text)
    mm = re.search(r'(\d+)m', text)
    sm = re.search(r'(\d+)s', text)
    if hm: h = int(hm.group(1))
    if mm: m = int(mm.group(1))
    if sm: s = int(sm.group(1))
    return h * 3600 + m * 60 + s


def extract_field(text: str, field: str, as_float: bool = False):
    """Estrae un campo dal testo di un'operazione (es. 'Cutting Distance: 1234.5mm')."""
    pattern = rf'{field}:\s*([\d.,]+)'
    match = re.search(pattern, text)
    if match:
        val = match.group(1).replace(",", "")
        return float(val) if as_float else val
    return None


def detect_strategy(op_text: str) -> str:
    """Determina la strategia CAM di un'operazione."""
    # Strategia esplicita
    strat_match = re.search(r'Strategy:\s*([A-Za-z]+(?:\s+[A-Za-z0-9]+)?)', op_text)
    if strat_match:
        raw = strat_match.group(1).strip()
        # Il parser a volte cattura "Adaptive Minimum" — troncare dopo la strategia
        known = ["Adaptive", "Facing", "Contour 2D", "Contour", "Drilling",
                 "Scallop", "Bore", "Pocket", "Slot", "Trace", "Radial",
                 "Spiral", "Morphed Spiral", "Parallel", "Pencil", "Steep and Shallow"]
        for k in known:
            if raw.startswith(k):
                return k
        return raw.split()[0]  # fallback: prima parola

    # Strategia implicita da Description
    desc_match = re.search(r'Description:\s*(?:\d+\s+)?(\w+)', op_text)
    if desc_match:
        desc_word = desc_match.group(1)
        if desc_word.lower().startswith("flat"):
            return "Flat"

    return "Unknown"


def extract_product_code(op_text: str) -> str:
    """Estrae il codice Product dall'operazione."""
    # Pattern: "Product: <codice>" fino a fine riga o prossimo campo
    match = re.search(r'Product:\s*(.+?)(?:\n|$)', op_text)
    if match:
        product = match.group(1).strip()
        # Pulizia: rimuovi eventuali campi successivi inline
        product = re.split(r'\s{2,}', product)[0].strip()
        # Rimuovi "fresa a punta tonda" prefix se presente (es. TP sheets)
        product = re.sub(r'^fresa a punta tonda\s*', '', product, flags=re.IGNORECASE)
        # Rimuovi " con inserto..." suffix
        product = re.split(r'\s+con\s+inserto', product, flags=re.IGNORECASE)[0].strip()
        return product
    return "N/A"


def parse_pdf(pdf_path: str) -> dict:
    """
    Parsa un PDF di operation sheet e restituisce i dati strutturati.

    Returns:
        {
            'name': str,               # Nome del documento
            'setups': [                 # Lista di setup (tipicamente 2)
                {
                    'program': str,
                    'cycle_time_s': int,
                    'n_operations': int,
                    'n_tools': int,
                    'operations': [     # Lista operazioni dettagliate
                        {
                            'op_num': int,
                            'op_total': int,
                            'description': str,
                            'strategy': str,
                            'tool_t': str,
                            'product': str,
                            'cutting_dist': float,
                            'rapid_dist': float,
                            'max_feedrate': float,
                            'cycle_time_s': int,
                        }, ...
                    ]
                }, ...
            ]
        }
    """
    result = {'name': '', 'setups': []}

    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    # Nome documento
    doc_match = re.search(r'Document Path:\s*(.+)', full_text)
    if doc_match:
        result['name'] = doc_match.group(1).strip()
    else:
        result['name'] = Path(pdf_path).stem

    # Dividi per Setup Sheet
    setup_blocks = re.split(r'(?=Setup Sheet for Program \d+)', full_text)
    setup_blocks = [b for b in setup_blocks if b.strip() and 'Setup Sheet for Program' in b]

    for block in setup_blocks:
        setup = {'program': '', 'cycle_time_s': 0, 'n_operations': 0, 'n_tools': 0, 'operations': []}

        # Program number
        prog_match = re.search(r'Setup Sheet for Program (\d+)', block)
        if prog_match:
            setup['program'] = prog_match.group(1)

        # Summary data from header
        nops_match = re.search(r'Number Of Operations:\s*(\d+)', block)
        if nops_match:
            setup['n_operations'] = int(nops_match.group(1))

        ntools_match = re.search(r'Number Of Tools:\s*(\d+)', block)
        if ntools_match:
            setup['n_tools'] = int(ntools_match.group(1))

        ct_match = re.search(r'Estimated Cycle Time:\s*([\dhms:]+)', block)
        if ct_match:
            setup['cycle_time_s'] = parse_cycle_time(ct_match.group(1))

        # Estrai operazioni individuali
        op_pattern = r'(Operation\s+(\d+)/(\d+)\s+(T\d+)\s+D\d+\s+L\d+.*?)(?=Operation\s+\d+/\d+|$)'
        ops = re.findall(op_pattern, block, re.DOTALL)

        for op_text, op_num, op_total, tool_t in ops:
            cutting = extract_field(op_text, 'Cutting Distance', as_float=True) or 0.0
            rapid = extract_field(op_text, 'Rapid Distance', as_float=True) or 0.0
            feedrate = extract_field(op_text, 'Maximum Feedrate', as_float=True) or 0.0

            # Cycle time dell'operazione
            op_ct_match = re.search(r'Estimated Cycle Time:\s*([\dhms:]+(?:\s*\([^)]*\))?)', op_text)
            op_ct = parse_cycle_time(op_ct_match.group(1)) if op_ct_match else 0

            # Description
            desc_match = re.search(r'Description:\s*(.+?)(?:\s{2,}|Maximum|Minimum|$)', op_text)
            description = desc_match.group(1).strip() if desc_match else ""

            strategy = detect_strategy(op_text)
            product = extract_product_code(op_text)

            setup['operations'].append({
                'op_num': int(op_num),
                'op_total': int(op_total),
                'description': description,
                'strategy': strategy,
                'tool_t': tool_t,
                'product': product,
                'cutting_dist': cutting,
                'rapid_dist': rapid,
                'max_feedrate': feedrate,
                'cycle_time_s': op_ct,
            })

        result['setups'].append(setup)

    return result


# ═══════════════════════════════════════════════════════════════════
# 2. CALCOLO METRICHE
# ═══════════════════════════════════════════════════════════════════

def extract_short_name(full_name: str) -> str:
    """Estrae un nome breve dal Document Path (es. 'NC02' da 'X_NC02-FORI_EDIT_12100709 v4')."""
    # Cerca pattern NC01, NC02, TP01, TP02, ecc.
    match = re.search(r'((?:NC|TP|GR)\d+)', full_name, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    # Fallback: prime parole significative
    clean = re.sub(r'[_\-]', ' ', full_name).split()
    return clean[0] if clean else full_name


def compute_metrics(parsed: dict, tool_life_s: int = 1200) -> dict:
    """Calcola tutte le metriche da un PDF parsato."""
    all_ops = []
    for setup in parsed['setups']:
        all_ops.extend(setup['operations'])

    if not all_ops:
        sys.exit(f"Errore: nessuna operazione trovata in '{parsed['name']}'")

    # Usa tempi ciclo dall'header del setup (includono cambi utensile e overhead)
    total_time = sum(s['cycle_time_s'] for s in parsed['setups'])
    # Fallback: se header non disponibile, usa somma operazioni
    if total_time == 0:
        total_time = sum(o['cycle_time_s'] for o in all_ops)

    total_cut = sum(o['cutting_dist'] for o in all_ops)
    total_rapid = sum(o['rapid_dist'] for o in all_ops)
    n_ops = len(all_ops)

    # Per setup — usa tempi header
    setup_times = [s['cycle_time_s'] for s in parsed['setups']]

    # Utensili univoci per Product
    products = set(o['product'] for o in all_ops if o['product'] != 'N/A')

    # Cambi utensile
    tool_changes = 0
    for setup in parsed['setups']:
        ops = setup['operations']
        for i in range(1, len(ops)):
            if ops[i]['tool_t'] != ops[i - 1]['tool_t']:
                tool_changes += 1

    # Strategie
    strategies = set(o['strategy'] for o in all_ops)
    strat_time = defaultdict(int)
    strat_count = defaultdict(int)
    for o in all_ops:
        strat_time[o['strategy']] += o['cycle_time_s']
        strat_count[o['strategy']] += 1

    # Tempo per utensile (per Product)
    tool_time = defaultdict(int)
    tool_trefs = defaultdict(set)
    for o in all_ops:
        tool_time[o['product']] += o['cycle_time_s']
        tool_trefs[o['product']].add(o['tool_t'])

    # Feedrate medio ponderato
    weighted_feed = sum(o['max_feedrate'] * o['cutting_dist'] for o in all_ops) / total_cut if total_cut else 0

    # Vita utile
    max_tool_time = max(tool_time.values()) if tool_time else 0
    max_tool_prod = max(tool_time, key=tool_time.get) if tool_time else "N/A"
    tools_over_50 = sum(1 for t in tool_time.values() if t / tool_life_s > 0.5)
    tools_over_75 = sum(1 for t in tool_time.values() if t / tool_life_s > 0.75)
    tools_over_100 = sum(1 for t in tool_time.values() if t / tool_life_s > 1.0)
    avg_util = sum(t / tool_life_s for t in tool_time.values()) / len(tool_time) if tool_time else 0

    n_products = len(products)

    short_name = extract_short_name(parsed['name'])

    return {
        'group': short_name,
        'full_name': parsed['name'],
        'total_time': total_time,
        'setup_times': setup_times,
        'total_cut': total_cut,
        'total_rapid': total_rapid,
        'n_ops': n_ops,
        'n_ops_per_setup': [len(s['operations']) for s in parsed['setups']],
        'n_products': n_products,
        'tc_total': tool_changes,
        'n_strategies': len(strategies),
        'strategies': strategies,
        'strat_time': dict(strat_time),
        'strat_count': dict(strat_count),
        'tool_time': dict(tool_time),
        'tool_trefs': {k: sorted(v) for k, v in tool_trefs.items()},
        'weighted_feed': weighted_feed,
        'max_tool_time': max_tool_time,
        'max_tool_prod': max_tool_prod,
        'tools_over_50': tools_over_50,
        'tools_over_75': tools_over_75,
        'tools_over_100': tools_over_100,
        'avg_util': avg_util,
        'cut_ratio': total_cut / (total_cut + total_rapid) if (total_cut + total_rapid) else 0,
        'ops_per_tool': n_ops / n_products if n_products else 0,
        'productivity': total_cut / (total_time / 60) if total_time else 0,
        'max_tool_pct_cycle': max_tool_time / total_time if total_time else 0,
        'tool_life_s': tool_life_s,
    }


# ═══════════════════════════════════════════════════════════════════
# 3. SISTEMA DI SCORING (VENDOR RATING)
# ═══════════════════════════════════════════════════════════════════

CATEGORY_WEIGHTS = {
    'Efficienza Temporale': 0.30,
    'Utilizzo Utensili': 0.20,
    'Vita Utile': 0.20,
    'Efficienza di Percorso': 0.15,
    'Complessità del Ciclo': 0.10,
    'Aggressività di Taglio': 0.05,
}


def relative_score(val_a: float, val_b: float, lower_is_better: bool = True):
    """Punteggio relativo: il migliore prende 100, l'altro in proporzione."""
    if val_a == 0 and val_b == 0:
        return 100.0, 100.0
    if lower_is_better:
        best = min(val_a, val_b)
        if val_a == 0: return 100.0, 0.0
        if val_b == 0: return 0.0, 100.0
        return round(best / val_a * 100, 1), round(best / val_b * 100, 1)
    else:
        best = max(val_a, val_b)
        if best == 0: return 100.0, 100.0
        return round(val_a / best * 100, 1), round(val_b / best * 100, 1)


def tool_life_score(metrics: dict) -> float:
    """
    Score non lineare per vita utile:
        ≤50% → 100 | 50–75% → 80 | 75–100% → 60 | >100% → penalità rapida
    """
    limit = metrics['tool_life_s']
    scores = []
    for t in metrics['tool_time'].values():
        pct = t / limit
        if pct <= 0.5:
            s = 100
        elif pct <= 0.75:
            s = 80
        elif pct <= 1.0:
            s = 60
        else:
            s = max(0, 60 - (pct - 1.0) * 200)
        scores.append(s)
    return round(sum(scores) / len(scores), 1) if scores else 100


def compute_scores(ma: dict, mb: dict):
    """
    Calcola punteggi per tutti i driver e per categoria.

    Returns:
        drivers: lista di tuple (categoria, nome_driver, raw_a, raw_b, score_a, score_b, disp_a, disp_b)
        cat_scores_a/b: dict {categoria: score}
        total_a/b: float punteggio pesato finale
    """
    drivers = []

    def add(cat, name, va, vb, sa, sb, da, db):
        drivers.append((cat, name, va, vb, sa, sb, da, db))

    # 1. EFFICIENZA TEMPORALE
    s1a, s1b = relative_score(ma['total_time'], mb['total_time'], True)
    add('Efficienza Temporale', 'Tempo ciclo complessivo',
        ma['total_time'], mb['total_time'], s1a, s1b,
        fmt_time(ma['total_time']), fmt_time(mb['total_time']))

    tma = ma['total_time'] / ma['n_ops'] if ma['n_ops'] else 0
    tmb = mb['total_time'] / mb['n_ops'] if mb['n_ops'] else 0
    s2a, s2b = relative_score(tma, tmb, True)
    add('Efficienza Temporale', 'Tempo medio per operazione',
        tma, tmb, s2a, s2b, fmt_time(tma), fmt_time(tmb))

    # 2. UTILIZZO UTENSILI
    s3a, s3b = relative_score(ma['n_products'], mb['n_products'], True)
    add('Utilizzo Utensili', 'N° utensili univoci',
        ma['n_products'], mb['n_products'], s3a, s3b,
        str(ma['n_products']), str(mb['n_products']))

    s4a, s4b = relative_score(ma['tc_total'], mb['tc_total'], True)
    add('Utilizzo Utensili', 'N° cambi utensile',
        ma['tc_total'], mb['tc_total'], s4a, s4b,
        str(ma['tc_total']), str(mb['tc_total']))

    # 3. VITA UTILE
    tls_a, tls_b = tool_life_score(ma), tool_life_score(mb)
    add('Vita Utile', 'Score vita utile (non lineare)',
        tls_a, tls_b, tls_a, tls_b,
        f"{tls_a:.1f}/100", f"{tls_b:.1f}/100")

    s6a, s6b = relative_score(ma['max_tool_pct_cycle'], mb['max_tool_pct_cycle'], True)
    add('Vita Utile', 'Concentrazione utensile più impiegato',
        ma['max_tool_pct_cycle'], mb['max_tool_pct_cycle'], s6a, s6b,
        f"{ma['max_tool_pct_cycle'] * 100:.1f}%", f"{mb['max_tool_pct_cycle'] * 100:.1f}%")

    pen_a = max(0, 100 - ma['tools_over_100'] * 50)
    pen_b = max(0, 100 - mb['tools_over_100'] * 50)
    add('Vita Utile', 'Penalità superamento vita (−50pt/utensile)',
        ma['tools_over_100'], mb['tools_over_100'], pen_a, pen_b,
        f"{ma['tools_over_100']} utensili", f"{mb['tools_over_100']} utensili")

    # 4. EFFICIENZA DI PERCORSO
    s8a, s8b = relative_score(ma['cut_ratio'], mb['cut_ratio'], False)
    add('Efficienza di Percorso', 'Rapporto taglio / (taglio + rapido)',
        ma['cut_ratio'], mb['cut_ratio'], s8a, s8b,
        f"{ma['cut_ratio'] * 100:.1f}%", f"{mb['cut_ratio'] * 100:.1f}%")

    da = ma['total_cut'] + ma['total_rapid']
    db = mb['total_cut'] + mb['total_rapid']
    s9a, s9b = relative_score(da, db, True)
    add('Efficienza di Percorso', 'Distanza complessiva',
        da, db, s9a, s9b, f"{da:.0f} mm", f"{db:.0f} mm")

    # 5. COMPLESSITA'
    s10a, s10b = relative_score(ma['n_ops'], mb['n_ops'], True)
    add('Complessità del Ciclo', 'N° operazioni totali',
        ma['n_ops'], mb['n_ops'], s10a, s10b,
        str(ma['n_ops']), str(mb['n_ops']))

    s11a, s11b = relative_score(ma['ops_per_tool'], mb['ops_per_tool'], True)
    add('Complessità del Ciclo', 'Rapporto operazioni / utensile',
        ma['ops_per_tool'], mb['ops_per_tool'], s11a, s11b,
        f"{ma['ops_per_tool']:.1f}", f"{mb['ops_per_tool']:.1f}")

    # 6. AGGRESSIVITA' DI TAGLIO
    s12a, s12b = relative_score(ma['weighted_feed'], mb['weighted_feed'], False)
    add('Aggressività di Taglio', 'Feedrate medio ponderato',
        ma['weighted_feed'], mb['weighted_feed'], s12a, s12b,
        f"{ma['weighted_feed']:.0f} mm/min", f"{mb['weighted_feed']:.0f} mm/min")

    s13a, s13b = relative_score(ma['productivity'], mb['productivity'], False)
    add('Aggressività di Taglio', 'Produttività [mm taglio / min ciclo]',
        ma['productivity'], mb['productivity'], s13a, s13b,
        f"{ma['productivity']:.0f}", f"{mb['productivity']:.0f}")

    # Calcolo punteggi categoria
    cat_scores_a, cat_scores_b = {}, {}
    for cat in CATEGORY_WEIGHTS:
        cd = [d for d in drivers if d[0] == cat]
        if cd:
            cat_scores_a[cat] = round(sum(d[4] for d in cd) / len(cd), 1)
            cat_scores_b[cat] = round(sum(d[5] for d in cd) / len(cd), 1)

    total_a = round(sum(cat_scores_a[c] * w for c, w in CATEGORY_WEIGHTS.items()), 1)
    total_b = round(sum(cat_scores_b[c] * w for c, w in CATEGORY_WEIGHTS.items()), 1)

    return drivers, cat_scores_a, cat_scores_b, total_a, total_b


# ═══════════════════════════════════════════════════════════════════
# 4. FORMATTAZIONE OUTPUT
# ═══════════════════════════════════════════════════════════════════

def fmt_time(s: float) -> str:
    s = int(s)
    m, sec = divmod(s, 60)
    h, m = divmod(m, 60)
    return f"{h}h {m:02d}m {sec:02d}s" if h > 0 else f"{m}m {sec:02d}s"


def print_report(ma, mb, drivers, csa, csb, ta, tb):
    """Stampa il report di benchmark su console."""
    na, nb = ma['group'], mb['group']
    fna, fnb = ma.get('full_name', na), mb.get('full_name', nb)
    W = 88

    print("\n" + "═" * W)
    print(f"{'VENDOR RATING — BENCHMARK CNC':^{W}}")
    print(f"{na}  vs  {nb}".center(W))
    print(f"({fna})".center(W))
    print(f"({fnb})".center(W))
    print("═" * W)

    # Score complessivo
    winner = na if ta > tb else (nb if tb > ta else "PARITÀ")
    print(f"\n  {'PUNTEGGIO FINALE':^30}  {na:>12}  {nb:>12}  {'Migliore':>10}")
    print(f"  {'─' * 30}  {'─' * 12}  {'─' * 12}  {'─' * 10}")
    print(f"  {'Score Complessivo':^30}  {ta:>11.1f}  {tb:>11.1f}  {winner:>10}")

    # Dettaglio per categoria
    print(f"\n  {'CATEGORIA':<32} {'Peso':>6}  {na:>10}  {nb:>10}  {'Migliore':>10}")
    print(f"  {'─' * 32} {'─' * 6}  {'─' * 10}  {'─' * 10}  {'─' * 10}")
    for cat, weight in CATEGORY_WEIGHTS.items():
        sa, sb = csa[cat], csb[cat]
        best = na if sa > sb else (nb if sb > sa else "=")
        marker_a = " ◄" if sa > sb else ""
        marker_b = " ◄" if sb > sa else ""
        print(f"  {cat:<32} {weight * 100:>5.0f}%  {sa:>8.1f}{marker_a:<2}  {sb:>8.1f}{marker_b:<2}  {best:>10}")

    # Dettaglio driver
    print(f"\n{'─' * W}")
    print(f"  {'DRIVER':<44} {na:>12}  {nb:>12}  {'Score ' + na:>10}  {'Score ' + nb:>10}")
    print(f"  {'─' * 44} {'─' * 12}  {'─' * 12}  {'─' * 10}  {'─' * 10}")

    current_cat = ""
    for cat, driver_name, raw_a, raw_b, score_a, score_b, disp_a, disp_b in drivers:
        if cat != current_cat:
            print(f"\n  ▸ {cat} ({CATEGORY_WEIGHTS[cat] * 100:.0f}%)")
            current_cat = cat
        marker_a = " ◄" if score_a > score_b else ""
        marker_b = " ◄" if score_b > score_a else ""
        print(f"    {driver_name:<42} {disp_a:>12}  {disp_b:>12}  {score_a:>8.1f}{marker_a:<2}  {score_b:>8.1f}{marker_b:<2}")

    # Allarmi vita utile
    limit = ma['tool_life_s']
    alarms_a = {p: t for p, t in ma['tool_time'].items() if t > limit}
    alarms_b = {p: t for p, t in mb['tool_time'].items() if t > limit}
    if alarms_a or alarms_b:
        print(f"\n{'─' * W}")
        print(f"  ⚠  ALLARMI VITA UTILE (soglia {limit // 60} min)")
        for p, t in alarms_a.items():
            trefs = ", ".join(ma['tool_trefs'].get(p, []))
            print(f"    [{na}]  {p} ({trefs}): {fmt_time(t)} = {t / limit * 100:.1f}% vita")
        for p, t in alarms_b.items():
            trefs = ", ".join(mb['tool_trefs'].get(p, []))
            print(f"    [{nb}]  {p} ({trefs}): {fmt_time(t)} = {t / limit * 100:.1f}% vita")

    # Metodologia
    print(f"\n{'─' * W}")
    print("  METODOLOGIA")
    print("  • Punteggio relativo: migliore = 100, altro proporzionale")
    print("  • Vita utile non lineare: ≤50%→100 | 50–75%→80 | 75–100%→60 | >100%→penalità rapida")
    print("  • Penalità assoluta: −50pt per ogni utensile oltre il 100% vita utile")
    print("  • Pesi: Tempo 30% | Utensili 20% | Vita 20% | Percorso 15% | Complessità 10% | Taglio 5%")
    print("═" * W + "\n")


# ═══════════════════════════════════════════════════════════════════
# 5. ESPORTAZIONE EXCEL (opzionale)
# ═══════════════════════════════════════════════════════════════════

def export_xlsx(ma, mb, drivers, csa, csb, ta, tb, xlsx_path: str):
    """Esporta il benchmark completo in un file Excel formattato."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Attenzione: openpyxl non installato, export Excel saltato.")
        return

    na, nb = ma['group'], mb['group']
    fna, fnb = ma.get('full_name', na), mb.get('full_name', nb)
    wb = Workbook()

    hf = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hfill = PatternFill("solid", fgColor="2F5496")
    cat_font = Font(bold=True, name="Arial", size=10, color="2F5496")
    cat_fill = PatternFill("solid", fgColor="D6E4F0")
    df = Font(name="Arial", size=10)
    bf = Font(bold=True, name="Arial", size=10)
    bf12 = Font(bold=True, name="Arial", size=12)
    green_font = Font(bold=True, name="Arial", size=10, color="217346")
    red_font = Font(bold=True, name="Arial", size=10, color="C00000")
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    ca_ = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la_ = Alignment(horizontal='left', vertical='center', wrap_text=True)
    better_fill = PatternFill("solid", fgColor="E2EFDA")
    worse_fill = PatternFill("solid", fgColor="FCE4EC")
    gold_fill = PatternFill("solid", fgColor="FFD700")

    # --- SCORECARD ---
    ws = wb.active
    ws.title = "Scorecard"
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value=f"VENDOR RATING: {na} vs {nb}").font = Font(bold=True, name="Arial", size=14, color="2F5496")
    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"{fna}  vs  {fnb}").font = Font(name="Arial", size=9, italic=True, color="666666")

    # Scores banner
    ws.cell(row=3, column=1, value=na).font = bf12; ws.cell(row=3, column=1).alignment = ca_
    c2 = ws.cell(row=3, column=2, value=f"{ta:.1f} / 100")
    c2.font = Font(bold=True, name="Arial", size=14, color="217346" if ta >= tb else "C00000"); c2.alignment = ca_
    if ta >= tb: c2.fill = gold_fill
    ws.cell(row=3, column=4, value=nb).font = bf12; ws.cell(row=3, column=4).alignment = ca_
    c5 = ws.cell(row=3, column=5, value=f"{tb:.1f} / 100")
    c5.font = Font(bold=True, name="Arial", size=14, color="217346" if tb >= ta else "C00000"); c5.alignment = ca_
    if tb >= ta: c5.fill = gold_fill

    # Detail headers
    row = 5
    for c, h in enumerate(["Categoria", "Driver", f"Valore {na}", f"Valore {nb}", f"Score {na}", f"Score {nb}", "Δ", "Migliore"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca_; cell.border = thin
    row += 1

    current_cat = ""
    for cat, dn, ra, rb, sa, sb, da, db in drivers:
        if cat != current_cat:
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = cat_fill; ws.cell(row=row, column=c).border = thin
            ws.cell(row=row, column=1, value=f"{cat} ({CATEGORY_WEIGHTS[cat] * 100:.0f}%)").font = cat_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1).fill = cat_fill
            current_cat = cat; row += 1

        ws.cell(row=row, column=2, value=dn).font = df; ws.cell(row=row, column=2).alignment = la_
        ws.cell(row=row, column=3, value=da).font = df; ws.cell(row=row, column=3).alignment = ca_
        ws.cell(row=row, column=4, value=db).font = df; ws.cell(row=row, column=4).alignment = ca_
        ws.cell(row=row, column=5, value=f"{sa:.1f}").font = bf; ws.cell(row=row, column=5).alignment = ca_
        ws.cell(row=row, column=6, value=f"{sb:.1f}").font = bf; ws.cell(row=row, column=6).alignment = ca_
        ws.cell(row=row, column=7, value=f"{abs(sa - sb):.1f}").font = df; ws.cell(row=row, column=7).alignment = ca_
        w = na if sa > sb else (nb if sb > sa else "=")
        ws.cell(row=row, column=8, value=w).font = green_font if w != "=" else df; ws.cell(row=row, column=8).alignment = ca_
        if sa > sb: ws.cell(row=row, column=5).fill = better_fill
        elif sb > sa: ws.cell(row=row, column=6).fill = better_fill
        for c in range(1, 9): ws.cell(row=row, column=c).border = thin
        row += 1

    # Category summary
    row += 1
    ws.cell(row=row, column=1, value="RIEPILOGO PER CATEGORIA").font = Font(bold=True, name="Arial", size=11, color="2F5496")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8); row += 1
    for c, h in enumerate(["Categoria", "Peso", f"Score {na}", f"Score {nb}", f"Pesato {na}", f"Pesato {nb}", "Δ", "Migliore"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca_; cell.border = thin
    row += 1
    for cat, weight in CATEGORY_WEIGHTS.items():
        sa, sb = csa[cat], csb[cat]
        ws.cell(row=row, column=1, value=cat).font = bf; ws.cell(row=row, column=1).alignment = la_
        ws.cell(row=row, column=2, value=f"{weight * 100:.0f}%").font = df; ws.cell(row=row, column=2).alignment = ca_
        ws.cell(row=row, column=3, value=f"{sa:.1f}").font = bf; ws.cell(row=row, column=3).alignment = ca_
        ws.cell(row=row, column=4, value=f"{sb:.1f}").font = bf; ws.cell(row=row, column=4).alignment = ca_
        ws.cell(row=row, column=5, value=f"{sa * weight:.1f}").font = df; ws.cell(row=row, column=5).alignment = ca_
        ws.cell(row=row, column=6, value=f"{sb * weight:.1f}").font = df; ws.cell(row=row, column=6).alignment = ca_
        ws.cell(row=row, column=7, value=f"{abs(sa - sb):.1f}").font = df; ws.cell(row=row, column=7).alignment = ca_
        w = na if sa > sb else (nb if sb > sa else "=")
        ws.cell(row=row, column=8, value=w).font = green_font if w != "=" else df; ws.cell(row=row, column=8).alignment = ca_
        if sa > sb: ws.cell(row=row, column=3).fill = better_fill
        elif sb > sa: ws.cell(row=row, column=4).fill = better_fill
        for c in range(1, 9): ws.cell(row=row, column=c).border = thin
        row += 1

    # Total row
    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = cat_fill; ws.cell(row=row, column=c).border = thin
    ws.cell(row=row, column=1, value="TOTALE").font = bf12
    ws.cell(row=row, column=2, value="100%").font = bf; ws.cell(row=row, column=2).alignment = ca_
    ws.cell(row=row, column=5, value=f"{ta:.1f}").font = bf12; ws.cell(row=row, column=5).alignment = ca_
    ws.cell(row=row, column=5).fill = gold_fill if ta >= tb else cat_fill
    ws.cell(row=row, column=6, value=f"{tb:.1f}").font = bf12; ws.cell(row=row, column=6).alignment = ca_
    ws.cell(row=row, column=6).fill = gold_fill if tb >= ta else cat_fill
    ws.cell(row=row, column=8, value=na if ta > tb else nb).font = Font(bold=True, name="Arial", size=12, color="217346")
    ws.cell(row=row, column=8).alignment = ca_

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 44
    for l in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[l].width = 16

    # --- VITA UTILE ---
    ws2 = wb.create_sheet("Vita Utile")
    all_prods = sorted(set(list(ma['tool_time'].keys()) + list(mb['tool_time'].keys())))
    for c, h in enumerate(["#", "Codice PRODUCT", "Rif. T", f"Tempo {na}", f"% Vita {na}",
                            f"Tempo {nb}", f"% Vita {nb}", f"Stato {na}", f"Stato {nb}"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca_; cell.border = thin

    limit = ma['tool_life_s']
    for i, prod in enumerate(all_prods, 1):
        r = i + 1
        ws2.cell(row=r, column=1, value=i).font = df; ws2.cell(row=r, column=1).alignment = ca_
        ws2.cell(row=r, column=2, value=prod).font = bf
        refs = set()
        if prod in ma['tool_trefs']: refs.update(ma['tool_trefs'][prod])
        if prod in mb['tool_trefs']: refs.update(mb['tool_trefs'][prod])
        ws2.cell(row=r, column=3, value=", ".join(sorted(refs))).font = Font(name="Arial", size=9)
        ws2.cell(row=r, column=3).alignment = ca_

        for col_t, col_p, col_s, metrics in [(4, 5, 8, ma), (6, 7, 9, mb)]:
            if prod in metrics['tool_time']:
                ts = metrics['tool_time'][prod]; pct = ts / limit
                ws2.cell(row=r, column=col_t, value=fmt_time(ts)).font = df; ws2.cell(row=r, column=col_t).alignment = ca_
                ws2.cell(row=r, column=col_p, value=f"{pct * 100:.1f}%").font = df; ws2.cell(row=r, column=col_p).alignment = ca_
                if pct > 1.0:
                    ws2.cell(row=r, column=col_s, value="⚠ SUPERATO").font = red_font
                    ws2.cell(row=r, column=col_p).fill = worse_fill; ws2.cell(row=r, column=col_p).font = red_font
                elif pct > 0.75:
                    ws2.cell(row=r, column=col_s, value="Attenzione").font = Font(name="Arial", size=10, color="FF8C00")
                    ws2.cell(row=r, column=col_p).fill = PatternFill("solid", fgColor="FFF2CC")
                elif pct > 0.5:
                    ws2.cell(row=r, column=col_s, value="Moderato").font = df
                else:
                    ws2.cell(row=r, column=col_s, value="OK").font = green_font
                    ws2.cell(row=r, column=col_p).fill = better_fill
                ws2.cell(row=r, column=col_s).alignment = ca_
            else:
                for cc in [col_t, col_p, col_s]:
                    ws2.cell(row=r, column=cc, value="—").font = Font(name="Arial", size=10, color="AAAAAA")
                    ws2.cell(row=r, column=cc).alignment = ca_
        for c in range(1, 10): ws2.cell(row=r, column=c).border = thin

    for i, w in enumerate([4, 28, 22, 14, 14, 14, 14, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(xlsx_path)
    print(f"\n  ✓ Report Excel salvato in: {xlsx_path}")


# ═══════════════════════════════════════════════════════════════════
# 6. MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="CNC Operation Sheet — Vendor Rating Benchmark",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempio:
  python benchmark_cnc.py  NC02_SHEET.pdf  TP02_SHEET.pdf
  python benchmark_cnc.py  NC02_SHEET.pdf  TP02_SHEET.pdf --xlsx report.xlsx
  python benchmark_cnc.py  NC02_SHEET.pdf  TP02_SHEET.pdf --tool-life 15
        """)
    parser.add_argument('pdf_a', help='PDF operation sheet del gruppo A')
    parser.add_argument('pdf_b', help='PDF operation sheet del gruppo B')
    parser.add_argument('--xlsx', help='Esporta risultati in file Excel', default=None)
    parser.add_argument('--tool-life', type=int, default=20,
                        help='Soglia vita utile utensile in minuti (default: 20)')

    args = parser.parse_args()
    tool_life_s = args.tool_life * 60

    # Parsing
    print(f"\n  Parsing {args.pdf_a} ...")
    parsed_a = parse_pdf(args.pdf_a)
    print(f"  → {parsed_a['name']}: {sum(len(s['operations']) for s in parsed_a['setups'])} operazioni in {len(parsed_a['setups'])} setup")

    print(f"  Parsing {args.pdf_b} ...")
    parsed_b = parse_pdf(args.pdf_b)
    print(f"  → {parsed_b['name']}: {sum(len(s['operations']) for s in parsed_b['setups'])} operazioni in {len(parsed_b['setups'])} setup")

    # Metriche
    ma = compute_metrics(parsed_a, tool_life_s)
    mb = compute_metrics(parsed_b, tool_life_s)

    # Scoring
    drivers, csa, csb, ta, tb = compute_scores(ma, mb)

    # Output
    print_report(ma, mb, drivers, csa, csb, ta, tb)

    # Excel export
    if args.xlsx:
        export_xlsx(ma, mb, drivers, csa, csb, ta, tb, args.xlsx)


if __name__ == '__main__':
    main()
